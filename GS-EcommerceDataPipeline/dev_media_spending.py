import pandas as pd

import glob

from collections import Counter

from dev_functions import*

import warnings
warnings.filterwarnings('ignore')

db='media_spending_raw'

### LAZ
def laz_aff(db,p,label):
    
    print('Pipeline : ' , label)
    
    print('')

    col_tmp=[]

    dataframe={}

    n=0

    print('### 1. Process files ')
    
    print('')
    
    for i in p:

        print('file : ', i)
        
        print('')
        
        df=pd.read_excel(i)
        
        col=[col for col in df.columns]
        
        date1=df['Date'].min()
        
        date1=date1[:10]
        
        date2=df['Date'].max()
        
        date2=date2[:10]
        
        state=compare_header_in_files(col_tmp,col)
        
        print('')
        
        if state==True:
            
            dataframe[n]=pd.concat([dataframe[n],df])
          
        else : 
           
            n+=1
           
            dataframe[n]=df
            
            warning(date1,date2)
            
            print('')
            
            col_tmp=col
    
    conclude(n)
    
    print('### 2. Pre-process dataframes and Storing to database.')
    
    print('')
    
    for key, df in dataframe.items():

        prepare_header_to_mysql(df)
        
        print('')
        
        df=df.sort_values('DATE').drop_duplicates()
        
        Date=store_data_to_local(df,label)
        
        print('')

        update_to_mysql(db,df,label,Date)
        
        print('')


def laz_search(db,p,label):
    
    print('Pipeline : ' , label)
    
    print('')

    col_tmp=[]

    dataframe={}

    n=0

    print('### 1. Process files ')
    
    print('')
    
    for i in p:

        print('file : ', i)
        
        print('')
        
        df=pd.read_excel(i)

        if df[df.columns[0]][4]=='Date':

            df=pd.read_excel(i, header=5, dtype=str)
        
        col=[col for col in df.columns]
        
        date1=df['Date'].min()
        
        date1=date1[:10]
        
        date2=df['Date'].max()
        
        date2=date2[:10]
        
        state=compare_header_in_files(col_tmp,col)
        
        print('')
        
        if state==True:
            
            dataframe[n]=pd.concat([dataframe[n],df])
          
        else : 
           
            n+=1
           
            dataframe[n]=df
            
            warning(date1,date2)
            
            print('')
            
            col_tmp=col
    
    conclude(n)
    
    print('### 2. Pre-process dataframes and Storing to database.')
    
    print('')
    
    for key, df in dataframe.items():

        prepare_header_to_mysql(df)
        
        print('')
        
        df=df.sort_values('DATE').drop_duplicates()
        
        Date=store_data_to_local(df,label)
        
        print('')

        update_to_mysql(db,df,label,Date)
        
        print('')


# def laz_cpas_fb(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')
    
#     col_tmp=[]

#     dataframe={}

#     n=0
    
#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:
        
#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_csv(i)
        
#         col=[col for col in df.columns]
        
#         df = df.rename(columns={"Reporting starts": "Date",})

#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]

#         state=compare_header_in_files(col,col_tmp)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
            
#         else : 
            
#             n+=1
            
#             dataframe[n]=df
            
#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col

#     conclude(5)
    
#     print('')
    
#     print('### 2. Pre-process dataframes and Storing to database.')
    
#     print('')
        
#     for key, df in dataframe.items():

#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         print('')
        
#         update_to_mysql(db,df,label,Date)
        
#         print('')


# def laz_cpas_msb_sku(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')
    
#     col_tmp=[]

#     dataframe={}

#     n=0
    
#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:
        
#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_excel(i, header=0, sheet_name='SKU Performance', dtype=str)
        
#         col=[col for col in df.columns]
    
#         df['Date']=[x[:4]+'-'+x[4:6]+'-'+x[6:] for x in df['Date'].tolist()]
        
#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]

#         state=compare_header_in_files(col,col_tmp)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
            
#         else : 
            
#             n+=1
            
#             dataframe[n]=df

#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col
    
#     conclude(n)
    
#     print('')
    
#     print('### 2. Pre-process dataframes and Storing to database.')

#     print('')

#     for key, df in dataframe.items():
        
#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         print('')

#         update_to_mysql(db,df,label,Date)
        
#         print('')


# def laz_cpas_msb_overview(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')

#     col_tmp=[]

#     dataframe={}

#     n=0
    
#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:
        
#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_excel(i, header=0, sheet_name='Overview', dtype=str)

#         col=[col for col in df.columns]

#         df['Date']=[x[:4]+'-'+x[4:6]+'-'+x[6:] for x in df['Date'].tolist()]
        
#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]
        
#         state=compare_header_in_files(col,col_tmp)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
            
#         else : 
            
#             n+=1
            
#             dataframe[n]=df

#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col
        
#     conclude(n)
    
#     print('')
    
#     print('### 2. Pre-process dataframes and Storing to database.')

#     print('')
    
#     for key, df in dataframe.items():
        
#         unidecode_header(df)
        
#         print('')
        
#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         print('')

#         update_to_mysql(db,df,label,Date)
        
#         print('')


###     SHP 
def shp_aff(db,p,label):
    
    print('Pipeline : ' , label)
    
    print('')

    col_tmp=[]

    dataframe={}

    n=0

    print('### 1. Process files ')
    
    print('')
    
    for i in p:

        print('file : ', i)
        
        print('')
        
        df=pd.read_excel(i)
        
        col=[col for col in df.columns]
        
        date=[x.split()[0].replace('/', '-') for x in df['Purchase Time'].tolist()]
        
        for n,x in enumerate(date):
            if len(x.split('-')[-1])==4:
                date[n]=x.split('-')[-1]+'-'+x.split('-')[1]+'-'+x.split('-')[0]

        for n,x in enumerate(date):
            a=x.split('-')
            month=a[1]
            datex=a[-1]
            if len(month)==1:
                month='0'+month
            if len(datex)==1:
                datex='0'+datex
            date[n]=a[0]+'-'+month+'-'+datex

        df['Purchase Time']=date
        
        df = df.rename(columns={'Purchase Time': "Date",})
        
        date1=df['Date'].min()
        
        date1=date1[:10]
        
        date2=df['Date'].max()
        
        date2=date2[:10]
        
        state=compare_header_in_files(col_tmp,col)
        
        print('')
        
        if state==True:
            
            dataframe[n]=pd.concat([dataframe[n],df])
          
        else : 
           
            n+=1
           
            dataframe[n]=df
            
            warning(date1,date2)
            
            print('')
            
            col_tmp=col
    
    conclude(n)
    
    print('### 2. Pre-process dataframes and Storing to database.')
    
    print('')
    
    for key, df in dataframe.items():

        prepare_header_to_mysql(df)
        
        print('')
        
        df=df.sort_values('DATE').drop_duplicates()
        
        Date=store_data_to_local(df,label)
        
        print('')

        update_to_mysql(db,df,label,Date)
        
        print('')


def shp_search(db,p,label):
    
    print('Pipeline : ' , label)
    
    print('')

    col_tmp=[]

    dataframe={}

    n=0

    print('### 1. Process files ')
    
    print('')
    
    for i in p:

        print('file : ', i)
        
        print('')
        
        df=pd.read_excel(i)
        
        col=[col for col in df.columns]
        
        date=i.split('B-')[-1].split('.xl')[0].replace('.','-')

        df['Date']=[date]*df.shape[0]
        
        date1=df['Date'].min()
        
        date1=date1[:10]
        
        date2=df['Date'].max()
        
        date2=date2[:10]
        
        state=compare_header_in_files(col_tmp,col)
        
        print('')
        
        if state==True:
            
            dataframe[n]=pd.concat([dataframe[n],df])
          
        else : 
           
            n+=1
           
            dataframe[n]=df
            
            warning(date1,date2)
            
            print('')
            
            col_tmp=col
    
    conclude(n)
    
    print('### 2. Pre-process dataframes and Storing to database.')
    
    print('')
    
    for key, df in dataframe.items():

        prepare_header_to_mysql(df)
        
        print('')
        
        df=df.sort_values('DATE').drop_duplicates()
        
        Date=store_data_to_local(df,label)
        
        print('')

        update_to_mysql(db,df,label,Date)
        
        print('')


# def shp_cpas_fb(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')

#     col_tmp=[]

#     dataframe={}

#     n=0

#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:

#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_excel(i)
        
#         col=[col for col in df.columns]
        
#         df = df.rename(columns={'Reporting starts': "Date",})
        
#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]
        
#         state=compare_header_in_files(col_tmp,col)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
          
#         else : 
           
#             n+=1
           
#             dataframe[n]=df
            
#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col
    
#     conclude(n)
    
#     print('### 2. Pre-process dataframes and Storing to database.')
    
#     print('')
    
#     for key, df in dataframe.items():

#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         print('')

#         update_to_mysql(db,df,label,Date)
        
#         print('')
        
        
# def shp_cpas_msb_sku(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')

#     col_tmp=[]

#     dataframe={}

#     n=0

#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:

#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_excel(i, sheet_name='By Day', header=0)

#         col=[col for col in df.columns]

#         df['Date']=[x[6:]+'-'+x[3:5]+'-'+x[0:2] for x in df['Date'].tolist()]
        
#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]
        
#         state=compare_header_in_files(col_tmp,col)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
          
#         else : 
           
#             n+=1
           
#             dataframe[n]=df
            
#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col
    
#     conclude(n)
    
#     print('### 2. Pre-process dataframes and Storing to database.')
    
#     print('')
    
#     for key, df in dataframe.items():

#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         print('')

#         update_to_mysql(db,df,label,Date)
        
#         print('')
      

# def shp_cpas_msb_camp(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')

#     col_tmp=[]

#     dataframe={}

#     n=0

#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:

#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_excel(i, sheet_name='Details Sheet', header=0,skiprows=[1])
        
#         col=[col for col in df.columns]
        
#         df = df.rename(columns={'Date(DD/MM/YYYY)': "Date",})
        
#         df['Date)']=[x.replace('/','') for x in df['Date']]

#         df['Date']=[x[4:8]+'-'+x[2:4]+'-'+x[0:2] for x in df['Date']]
        
#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]
        
#         state=compare_header_in_files(col_tmp,col)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
          
#         else : 
           
#             n+=1
           
#             dataframe[n]=df
            
#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col
    
#     conclude(n)
    
#     print('### 2. Pre-process dataframes and Storing to database.')
    
#     print('')
    
#     for key, df in dataframe.items():

#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         print('')

#         update_to_mysql(db,df,label,Date)
        
#         print('')


### TIKI 

def tiki_aff(db,p,label):
    
    from openpyxl import load_workbook
    
    print('Pipeline : ' , label)
    
    print('')

    col_tmp=[]

    dataframe={}

    n=0

    print('### 1. Process files ')
    
    print('')
    
    for i in p:

        print('file : ', i)
        
        print('')
    
        df=[]
        
        a=[]
        
        for i in p:
        
            wb = load_workbook(i, read_only=True, keep_links=False)
            
            for x in wb.sheetnames:
                
                y=pd.read_excel(i, sheet_name=x, dtype=str, skiprows=0)

                a.append(y)

        df=pd.concat(a)
        
        col=[col for col in df.columns]
        
        date1=df['Date'].min()
        
        date1=date1[:10]
        
        date2=df['Date'].max()
        
        date2=date2[:10]
        
        state=compare_header_in_files(col_tmp,col)
        
        print('')
        
        if state==True:
            
            dataframe[n]=pd.concat([dataframe[n],df])
          
        else : 
           
            n+=1
           
            dataframe[n]=df
            
            warning(date1,date2)
            
            print('')
            
            col_tmp=col
    
    conclude(n)
    
    print('### 2. Pre-process dataframes and Storing to database.')
    
    print('')
    
    for key, df in dataframe.items():

        prepare_header_to_mysql(df)
        
        print('')
        
        df=df.sort_values('DATE').drop_duplicates()
        
        Date=store_data_to_local(df,label)
        
        print('')

        update_to_mysql(db,df,label,Date)
        
        print('')


def tiki_search(db,p,label):
    
    from openpyxl import load_workbook
    
    print('Pipeline : ' , label)
    
    print('')

    col_tmp=[]

    dataframe={}

    n=0

    print('### 1. Process files ')
    
    print('')
    
    for i in p:

        print('file : ', i)
        
        print('')
        
        df=[]
        
        a=[]
        
        wb = load_workbook(i, read_only=True, keep_links=False)
        
        for x in wb.sheetnames:
            y=pd.read_excel(i, sheet_name=x, dtype=str, skiprows=0)

            a.append(y)

        df=pd.concat(a)
        
        col=[col for col in df.columns]
        
        date1=df['Date'].min()
        
        date1=date1[:10]
        
        date2=df['Date'].max()
        
        date2=date2[:10]
        
        state=compare_header_in_files(col_tmp,col)
        
        print('')
        
        if state==True:
            
            dataframe[n]=pd.concat([dataframe[n],df])
          
        else : 
           
            n+=1
           
            dataframe[n]=df
            
            warning(date1,date2)
            
            print('')
            
            col_tmp=col
    
    conclude(n)
    
    print('### 2. Pre-process dataframes and Storing to database.')
    
    print('')
    
    for key, df in dataframe.items():

        prepare_header_to_mysql(df)
        
        print('')
        
        df=df.sort_values('DATE').drop_duplicates()
        
        Date=store_data_to_local(df,label)
        
        print('')

        update_to_mysql(db,df,label,Date)
        
        print('')


# def tiki_cpas_fb(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')

#     col_tmp=[]

#     dataframe={}

#     n=0

#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:

#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_excel(i)
        
#         col=[col for col in df.columns]
        
#         df = df.rename(columns={'Reporting starts': "Date",})
        
#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]
        
#         state=compare_header_in_files(col_tmp,col)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
          
#         else : 
           
#             n+=1
           
#             dataframe[n]=df
            
#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col
    
#     conclude(n)
    
#     print('### 2. Pre-process dataframes and Storing to database.')
    
#     print('')
    
#     for key, df in dataframe.items():

#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         print('')

#         update_to_mysql(db,df,label,Date)
        
#         print('')
        
        
# def tiki_cpas_msb_camp(db,p,label):
    
#     print('Pipeline : ' , label)
    
#     print('')

#     col_tmp=[]

#     dataframe={}

#     n=0

#     print('### 1. Process files ')
    
#     print('')
    
#     for i in p:

#         print('file : ', i)
        
#         print('')
        
#         df=pd.read_excel(i,skiprows=2)
        
#         col=[col for col in df.columns]
       
#         df['Date Time']=[str(x).split()[0] for x in df['Date Time']]
        
#         df = df.rename(columns={'Date Time': "Date",})
 
#         date1=df['Date'].min()
        
#         date1=date1[:10]
        
#         date2=df['Date'].max()
        
#         date2=date2[:10]
        
#         state=compare_header_in_files(col_tmp,col)
        
#         print('')
        
#         if state==True:
            
#             dataframe[n]=pd.concat([dataframe[n],df])
          
#         else : 
           
#             n+=1
           
#             dataframe[n]=df
            
#             warning(date1,date2)
            
#             print('')
            
#             col_tmp=col
    
#     conclude(n)
    
#     print('### 2. Pre-process dataframes and Storing to database.')
    
#     print('')
    
#     for key, df in dataframe.items():

#         prepare_header_to_mysql(df)
        
#         print('')
        
#         df=df.sort_values('DATE').drop_duplicates()
        
#         Date=store_data_to_local(df,label)
        
#         Date=Date[:8]
        
#         print('')

#         update_to_mysql(db,df,label,Date)
        
#         print('')
      
